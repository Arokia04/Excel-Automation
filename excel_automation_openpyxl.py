import os
import openpyxl
              #Opening an Excel Document
excel = 'C:\\\\Users\\\\arajc\\\\Desktop\\\\sample.xlsx'
open_document = openpyxl.load_workbook(excel)
print(open_document)

              #Sheet Names
#print(open_document.get_sheet_names())

            #Accessing Cells using row and column combination
sheet = open_document.get_sheet_by_name('abc')
print(sheet['A2'].value)
           #Accessing Cells using row and column seperate
#print(sheet.cell(row = 4, column = 1).value)

           #Accessing a Range of Cells
#multiple_cells = sheet['A1':'B3']
#for row in multiple_cells:
   # for cell in row:
#print (cell.value)

            #Accessing all rows Number :1  columns

#mylist = [] 
#for row in sheet.iter_rows('A{}:A{}'.format(sheet.min_row,sheet.max_row)):
  #  for cell in row:
  #      mylist.append(cell.value)
#print (mylist)

        #Accessing all rows Number :1  rows
  
#for col in sheet.iter_cols(min_row=1, max_col=3, max_row=2):
  #  for cell in col:
 #       print(cell.value)
