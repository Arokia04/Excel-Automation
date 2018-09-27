from openpyxl import load_workbook
# The source xlsx file is named as source.xlsx
wb=load_workbook("sample.xlsx")
wb.save('sample1.xlsx')
ws = wb.active
first_column = ws['B']

# Print the contents
for x in range(len(first_column)): 
    print(first_column[x].value) 
